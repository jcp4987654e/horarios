import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import colorsys
import re

# --- Configuración de la página de Streamlit ---
st.set_page_config(
    page_title="Mi Horario Interactivo",
    page_icon="🗓️",
    layout="wide"
)

# --- DATOS DEL HORARIO (Ahora en Python) ---
schedule_data = {
    'Lunes': [ { 'start': '00:00', 'end': '05:59', 'task': 'Durmiendo' }, { 'start': '06:00', 'end': '06:29', 'task': 'Me despierto' }, { 'start': '06:30', 'end': '07:39', 'task': 'Preparándome' }, { 'start': '07:40', 'end': '07:59', 'task': 'Entro a clases' }, { 'start': '08:00', 'end': '14:59', 'task': 'En clase' }, { 'start': '15:00', 'end': '15:39', 'task': 'Salgo de clase' }, { 'start': '15:40', 'end': '15:59', 'task': 'Llego a casa' }, { 'start': '16:00', 'end': '16:59', 'task': 'Descanso' }, { 'start': '17:00', 'end': '17:59', 'task': 'Libre' }, { 'start': '18:00', 'end': '18:59', 'task': 'Ejercicio' }, { 'start': '19:00', 'end': '19:29', 'task': 'Neotech' }, { 'start': '19:30', 'end': '21:29', 'task': 'Marketing Digital' }, { 'start': '21:30', 'end': '22:59', 'task': 'Libre' }, { 'start': '23:00', 'end': '23:59', 'task': 'Durmiendo' } ],
    'Martes': [ { 'start': '00:00', 'end': '00:29', 'task': 'Apagar todo' }, { 'start': '00:30', 'end': '00:59', 'task': 'A dormir' }, { 'start': '01:00', 'end': '08:29', 'task': 'Durmiendo' }, { 'start': '08:30', 'end': '08:59', 'task': 'Me despierto' }, { 'start': '09:00', 'end': '10:59', 'task': 'METROGAS' }, { 'start': '11:00', 'end': '14:59', 'task': 'En clase' }, { 'start': '15:00', 'end': '15:39', 'task': 'Salgo de clase' }, { 'start': '15:40', 'end': '15:59', 'task': 'Llego a casa' }, { 'start': '16:00', 'end': '21:59', 'task': 'Libre' }, { 'start': '22:00', 'end': '22:29', 'task': 'Apagar todo' }, { 'start': '22:30', 'end': '22:59', 'task': 'A dormir' }, { 'start': '23:00', 'end': '23:59', 'task': 'Durmiendo' } ],
    'Miércoles': [ { 'start': '00:00', 'end': '05:59', 'task': 'Durmiendo' }, { 'start': '06:00', 'end': '06:29', 'task': 'Me despierto' }, { 'start': '06:30', 'end': '07:39', 'task': 'Preparándome' }, { 'start': '07:40', 'end': '07:59', 'task': 'Entro a clases' }, { 'start': '08:00', 'end': '14:59', 'task': 'En clase' }, { 'start': '15:00', 'end': '15:39', 'task': 'Salgo de clase' }, { 'start': '15:40', 'end': '15:59', 'task': 'Llego a casa' }, { 'start': '16:00', 'end': '16:59', 'task': 'Descanso' }, { 'start': '17:00', 'end': '17:59', 'task': 'Libre' }, { 'start': '18:00', 'end': '18:59', 'task': 'Ejercicio' }, { 'start': '19:00', 'end': '19:29', 'task': 'Neotech' }, { 'start': '19:30', 'end': '21:29', 'task': 'CS50x Harvard' }, { 'start': '21:30', 'end': '22:29', 'task': 'Libre' }, { 'start': '22:30', 'end': '22:59', 'task': 'A dormir' }, { 'start': '23:00', 'end': '23:59', 'task': 'Durmiendo' } ],
    'Jueves': [ { 'start': '00:00', 'end': '05:59', 'task': 'Durmiendo' }, { 'start': '06:00', 'end': '06:29', 'task': 'Me despierto' }, { 'start': '06:30', 'end': '07:39', 'task': 'Preparándome' }, { 'start': '07:40', 'end': '07:59', 'task': 'Entro a clases' }, { 'start': '08:00', 'end': '14:59', 'task': 'En clase' }, { 'start': '15:00', 'end': '15:39', 'task': 'Salgo de clase' }, { 'start': '15:40', 'end': '15:59', 'task': 'Llego a casa' }, { 'start': '16:00', 'end': '17:59', 'task': 'Libre' }, { 'start': '18:00', 'end': '18:59', 'task': 'Curso IA' }, { 'start': '19:00', 'end': '19:29', 'task': 'Neotech' }, { 'start': '19:30', 'end': '21:29', 'task': 'Convenio Multilateral' }, { 'start': '21:30', 'end': '23:59', 'task': 'Libre' } ],
    'Viernes': [ { 'start': '00:00', 'end': '00:29', 'task': 'Apagar todo' }, { 'start': '00:30', 'end': '00:59', 'task': 'A dormir' }, { 'start': '01:00', 'end': '08:29', 'task': 'Durmiendo' }, { 'start': '08:30', 'end': '08:59', 'task': 'Me despierto' }, { 'start': '09:00', 'end': '09:29', 'task': 'Preparándome' }, { 'start': '09:30', 'end': '09:59', 'task': 'Entro a clase' }, { 'start': '10:00', 'end': '14:59', 'task': 'En clase' }, { 'start': '15:00', 'end': '15:39', 'task': 'Salgo de clase' }, { 'start': '15:40', 'end': '15:59', 'task': 'Llego a casa' }, { 'start': '16:00', 'end': '18:59', 'task': 'Ejercicio' }, { 'start': '19:00', 'end': '19:29', 'task': 'Neotech' }, { 'start': '19:30', 'end': '21:29', 'task': 'Monotributo' }, { 'start': '21:30', 'end': '23:59', 'task': 'Libre' } ],
    'Sábado': [ { 'start': '00:00', 'end': '23:59', 'task': 'Libre' } ],
    'Domingo': [ { 'start': '00:00', 'end': '08:59', 'task': 'Libre' }, { 'start': '09:00', 'end': '16:59', 'task': 'Durmiendo' }, { 'start': '17:00', 'end': '18:59', 'task': 'Ejercicio' }, { 'start': '19:00', 'end': '21:29', 'task': 'Neotech' }, { 'start': '21:30', 'end': '22:29', 'task': 'Libre' }, { 'start': '22:30', 'end': '22:59', 'task': 'A dormir' }, { 'start': '23:00', 'end': '23:59', 'task': 'Durmiendo' } ]
}

# --- LÓGICA DE ESTILOS Y COLORES (Ahora en Python) ---
task_styles = {
    'Durmiendo': { 'emoji': '😴', 'color': 'hsl(220, 25%, 94%)' }, 'Libre': { 'emoji': '😎', 'color': 'hsl(140, 35%, 95%)' }, 'En clase': { 'emoji': '📚', 'color': 'hsl(45, 70%, 94%)' }, 'Preparándome': { 'emoji': '☕', 'color': 'hsl(30, 65%, 94%)' }, 'Llego a casa': { 'emoji': '🏡', 'color': 'hsl(180, 40%, 94%)' }, 'Salgo de clase': { 'emoji': '🚶‍♂️', 'color': 'hsl(200, 50%, 95%)' }, 'Descanso': { 'emoji': '🧘', 'color': 'hsl(150, 40%, 96%)' }, 'Ejercicio': { 'emoji': '💪', 'color': 'hsl(0, 60%, 95%)' }, 'Neotech': { 'emoji': '💻', 'color': 'hsl(240, 50%, 96%)' }, 'Marketing Digital': { 'emoji': '📈', 'color': 'hsl(260, 60%, 96%)' }, 'Convenio Multilateral': { 'emoji': '📄', 'color': 'hsl(280, 50%, 97%)' }, 'CS50x Harvard': { 'emoji': '�', 'color': 'hsl(300, 50%, 97%)' }, 'Monotributo': { 'emoji': '🧾', 'color': 'hsl(320, 50%, 97%)' }, 'Curso IA': { 'emoji': '🤖', 'color': 'hsl(340, 50%, 97%)' }, 'METROGAS': { 'emoji': '🔥', 'color': 'hsl(25, 60%, 95%)' }, 'Apagar todo': { 'emoji': '🔌', 'color': 'hsl(220, 20%, 90%)' }, 'A dormir': { 'emoji': '🛌', 'color': 'hsl(220, 25%, 92%)' }, 'Entro a clases': { 'emoji': '🏫', 'color': 'hsl(40, 70%, 93%)' },
}
generated_colors = {}

# --- FUNCIONES DE AYUDA ---

def time_to_minutes(time_str):
    """Convierte un string de tiempo 'HH:MM' a minutos totales."""
    hours, minutes = map(int, time_str.split(':'))
    return hours * 60 + minutes

def find_task_for_time(day, time_in_minutes):
    """Encuentra la tarea para un día y minuto específicos."""
    day_schedule = schedule_data.get(day, [])
    task_entry = next((entry for entry in day_schedule if time_to_minutes(entry['start']) <= time_in_minutes <= time_to_minutes(entry['end'])), None)
    return task_entry['task'] if task_entry else 'Libre'

def hsl_to_hex(h, s, l):
    """Convierte color HSL a código HEX para Excel."""
    try:
        # Convertir porcentajes a flotantes 0-1
        l /= 100.0
        s /= 100.0
        # Convertir HSL a RGB
        r, g, b = colorsys.hls_to_rgb(h / 360.0, l, s)
        # Convertir RGB 0-1 a 0-255 y luego a HEX
        return '{:02x}{:02x}{:02x}'.format(int(r * 255), int(g * 255), int(b * 255)).upper()
    except Exception:
        return 'FFFFFF' # Color blanco por defecto en caso de error

def get_task_style_py(task_name):
    """Obtiene el estilo (emoji y color) para una tarea. Genera uno si no existe."""
    if task_name in task_styles:
        return task_styles[task_name]
    if task_name not in generated_colors:
        h = sum(ord(c) for c in task_name) % 360
        generated_colors[task_name] = {'emoji': '✨', 'color': f'hsl({h}, 40%, 96%)'}
    return generated_colors[task_name]

# --- LÓGICA PARA GENERAR ARCHIVOS ---

@st.cache_data
def create_excel_file():
    """Crea un archivo Excel en memoria con el horario y los colores."""
    output = io.BytesIO()
    
    # Crear un DataFrame de pandas con el horario
    days = list(schedule_data.keys())
    times = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 30)]
    df = pd.DataFrame(index=times, columns=days)
    
    for day in days:
        for time_str in times:
            minutes = time_to_minutes(time_str)
            df.loc[time_str, day] = find_task_for_time(day, minutes)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Horario', index=True)
        workbook = writer.book
        worksheet = writer.sheets['Horario']

        # Aplicar colores a las celdas
        for r_idx, row in enumerate(worksheet.iter_rows(min_row=2, min_col=2)): # Empezar desde la celda B2
            for c_idx, cell in enumerate(row):
                task_name = cell.value
                if task_name:
                    style = get_task_style_py(task_name)
                    color_str = style['color']
                    # Extraer H, S, L del string 'hsl(H, S%, L%)'
                    match = re.search(r"hsl\((\d+),\s*([\d.]+)%,\s*([\d.]+)%\)", color_str)
                    if match:
                        h, s, l = map(float, match.groups())
                        hex_color = hsl_to_hex(h, s, l)
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    
    return output.getvalue()

def generate_html_table():
    """Genera una tabla HTML estática del horario para visualización."""
    days = list(schedule_data.keys())
    
    # Estilos CSS para la tabla
    html = """
    <style>
        body { font-family: 'Poppins', sans-serif; }
        table { border-collapse: collapse; width: 100%; font-size: 14px; }
        th, td { border: 1px solid #e2e8f0; text-align: center; padding: 12px; }
        thead th { background-color: #f1f5f9; position: sticky; top: 0; z-index: 1; }
        tbody td div { display: flex; align-items: center; justify-content: center; gap: 8px; }
        tbody td span:first-child { font-size: 1.2em; }
    </style>
    <table>
        <thead>
            <tr>
                <th>Hora</th>
    """
    for day in days:
        html += f"<th>{day}</th>"
    html += "</tr></thead><tbody>"

    # Generar filas de la tabla
    for total_minutes in range(0, 24 * 60, 30):
        hours = total_minutes // 60
        minutes = total_minutes % 60
        time_str = f"{hours:02d}:{minutes:02d}"
        html += f"<tr><td>{time_str}</td>"
        for day in days:
            task = find_task_for_time(day, total_minutes)
            style = get_task_style_py(task)
            color = style['color']
            emoji = style['emoji']
            html += f'<td style="background-color:{color};"><div><span>{emoji}</span><span>{task}</span></div></td>'
        html += "</tr>"

    html += "</tbody></table>"
    return html

# --- INTERFAZ DE STREAMLIT ---

st.title("🗓️ Visualizador de Horario")
st.write("Esta aplicación muestra tu horario personal. Puedes exportarlo a Excel desde la barra lateral.")

# Barra lateral con opciones
st.sidebar.header("Opciones de Exportación")
excel_data = create_excel_file()
st.sidebar.download_button(
    label="📥 Descargar como Excel",
    data=excel_data,
    file_name="mi_horario.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.sidebar.header("Instrucciones")
st.sidebar.markdown("""
1.  **Guarda este código** en un archivo (ej: `app.py`).
2.  Crea un archivo `requirements.txt` en la misma carpeta con el contenido del Canvas correspondiente.
3.  Instala las librerías:
    ```bash
    pip install -r requirements.txt
    ```
4.  Ejecuta la aplicación:
    ```bash
    streamlit run app.py
    ```
""")


# Mostrar la tabla HTML
# Nota: La edición en vivo ya no es posible ya que los datos se manejan en Python.
st.header("Vista Previa del Horario")
html_content = generate_html_table()
components.html(html_content, height=800, scrolling=True)
�
